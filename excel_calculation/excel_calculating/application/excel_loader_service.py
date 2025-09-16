from ..domain.column_statistics import SheetData
from pandas import ExcelFile, read_excel, notna, to_numeric
from ..domain.exceptions import ColumnNamesNotExsistingException
from typing import cast
from openpyxl import load_workbook, Workbook
from ..domain.column_statistics import ColumnDataType


def find_columns_and_row_in_sheets(
    columns_names: set[str], excel_file: ExcelFile
) -> list[SheetData]:
    results: list[SheetData] = []
    used_columns: list[str] = []
    for sheet in excel_file.sheet_names:
        preview = read_excel(excel_file, sheet_name=sheet, header=None, nrows=6)
        for row_idx in range(len(preview)):
            row_values = preview.iloc[row_idx].astype(str).str.strip().str.lower()
            matches = columns_names.intersection(row_values)
            if matches:
                used_columns.extend(matches)
                results.append(
                    SheetData(
                        sheet_name=cast(str, sheet),
                        header_row=row_idx,
                        columns_name=matches,
                    )
                )
                break
    if columns_names != set(used_columns):
        raise ColumnNamesNotExsistingException(columns_names.difference(used_columns))
    return results


def count_reduntand_rows(
    excel_file: ExcelFile, sheet_name: str, header_row, last_n: int = 10
) -> int:
    """
    This method counts rows that may contain only some additional calculations.
    args:
    excel_file: ExcelFile - file for work
    sheet_name: str - sheet in which we check table
    last_n: int - how many rows we want to check - default 10
    """
    wb = load_workbook(filename=excel_file, data_only=False)
    ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]
    last_rows = list(ws.iter_rows(values_only=False))[-last_n:]
    formula_rows_from_bottom = 0
    for row in reversed(last_rows):
        if all(
            str(cell.value).startswith("=") for cell in row if cell.value is not None
        ):
            formula_rows_from_bottom += 1
        else:
            break
    return formula_rows_from_bottom


def check_if_any_formula(row: list[str]) -> bool:
    return any(cell.startswith("=") for cell in row if cell is not None)


def check_if_any_value(row: list[str]) -> bool:
    return any(
        notna(to_numeric(cell, errors="coerce")) for cell in row if cell is not None
    )


def establish_column_type(
    excel_workbook: Workbook, column_index: int
) -> ColumnDataType:
    col_cells = [
        str(row[column_index].value)
        for row in excel_workbook.iter_rows()
        if row[column_index].value is not None
    ]
    if check_if_any_value(col_cells) and check_if_any_formula(col_cells):
        return ColumnDataType.MIX
    return ColumnDataType.ONE_TYPE
