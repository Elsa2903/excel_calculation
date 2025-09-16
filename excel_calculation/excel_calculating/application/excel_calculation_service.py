from pandas import ExcelFile, read_excel, to_numeric, DataFrame
from ..domain.column_statistics import (
    ColumnStatistics,
    ColumnDataType,
    CalculationResponse,
)
from .excel_loader_service import (
    find_columns_and_row_in_sheets,
    establish_column_type,
    count_reduntand_rows,
)
from openpyxl.utils import get_column_letter
from ..domain.exceptions import (
    CannotCalculateDataForThatColumnException,
    UploadTypeInWrongTypeException,
    CannotCalculateDataForAllColumnsException,
)
from typing import cast
from io import BufferedReader

from openpyxl import load_workbook, Workbook


def calculate_column_mix_statistics(
    excel_workbook: Workbook, df: DataFrame, column_index: int, header_row: int
) -> CalculationResponse:
    sum_value = 0.0
    length = 0
    for i in range(len(df)):
        cell = excel_workbook.cell(row=header_row + 1 + i + 1, column=column_index + 1)
        if isinstance(cell.value, (int, float)):
            sum_value += float(df.iloc[i, column_index])
            length += 1
    return CalculationResponse(
        sum=sum_value, avg=(sum_value / length) if length != 0 else 0.0
    )


def calculate_columns_statistics(
    columns_names: list[str], uploaded_file: BufferedReader
) -> list[ColumnStatistics]:
    try:
        excel_file = ExcelFile(uploaded_file)
    except Exception:
        raise UploadTypeInWrongTypeException(uploaded_file.name)
    sheets = find_columns_and_row_in_sheets(
        columns_names={col.strip().lower() for col in columns_names},
        excel_file=excel_file,
    )
    result = []
    for i in sheets:
        reduntant_last_rows_number = count_reduntand_rows(
            excel_file=uploaded_file, sheet_name=i.sheet_name, header_row=i.header_row
        )
        df = read_excel(
            excel_file,
            sheet_name=i.sheet_name,
            header=i.header_row,
            skipfooter=reduntant_last_rows_number,
        )
        df.columns = df.columns.str.strip().str.lower()
        col_names = {
            col
            for col in df.columns
            for col_name in i.columns_name
            if col.lower().startswith(col_name)
        }
        for col in col_names:
            column_index: int = cast(int, df.columns.get_loc(col))
            try:
                result.append(
                    ColumnStatistics(
                        sheet_column=get_column_letter(column_index + 1),
                        column=col,
                        sheet=i.sheet_name,
                        sum=round(df[col].sum(), 2),
                        avg=round(df[col].mean(), 2),
                    )
                )
            except TypeError as err:
                raise CannotCalculateDataForThatColumnException(col) from err

    return result


def calculate_columns_statistics_check_cell(
    columns_names: list[str], uploaded_file: BufferedReader
) -> list[ColumnStatistics]:
    try:
        excel_file = ExcelFile(uploaded_file)
    except Exception:
        raise UploadTypeInWrongTypeException(uploaded_file.name)

    sheets = find_columns_and_row_in_sheets(
        columns_names={col.strip().lower() for col in columns_names},
        excel_file=excel_file,
    )
    result = []
    for i in sheets:
        wb = load_workbook(filename=uploaded_file, data_only=False)
        ws = (
            wb[i.sheet_name]
            if isinstance(i.sheet_name, str)
            else wb.worksheets[i.sheet_name]
        )
        df = read_excel(
            excel_file,
            sheet_name=i.sheet_name,
            header=i.header_row,
        )
        column_names_mapping = {
            col_name.strip().lower(): col_name for col_name in df.columns
        }

        col_names = {
            col
            for col in column_names_mapping.keys()
            for col_name in i.columns_name
            if col.lower().startswith(col_name)
        }

        # check if all are numeric values
        if not all(
            df[[column_names_mapping[col] for col in col_names]].apply(
                lambda col: to_numeric(col.dropna(), errors="coerce").notna().all()
            )
        ):
            raise CannotCalculateDataForAllColumnsException
        for col in col_names:
            column_index: int = cast(int, df.columns.get_loc(column_names_mapping[col]))
            match establish_column_type(ws, column_index):
                case ColumnDataType.ONE_TYPE:
                    result.append(
                        ColumnStatistics(
                            sheet_column=get_column_letter(column_index + 1),
                            column=col,
                            sheet=i.sheet_name,
                            sum=round(df[column_names_mapping[col]].sum(), 2),
                            avg=round(df[column_names_mapping[col]].mean(), 2),
                        )
                    )

                case ColumnDataType.MIX:
                    response = calculate_column_mix_statistics(
                        ws, df, column_index, header_row=i.header_row
                    )
                    result.append(
                        ColumnStatistics(
                            sheet_column=get_column_letter(column_index + 1),
                            column=col,
                            sheet=i.sheet_name,
                            sum=round(response.sum, 2),
                            avg=round(response.avg, 2),
                        )
                    )
                case _:
                    raise CannotCalculateDataForThatColumnException(col)
    return result
